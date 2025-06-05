from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import csv
from io import StringIO, BytesIO
import os
import tempfile
import openpyxl
from odoo.exceptions import UserError
from odoo.tools import _
import logging

_logger = logging.getLogger(__name__)

class PropertyImportWizard(models.TransientModel):
    _name = 'property.import.wizard'
    _description = 'Wizard to import property data'

    data_file = fields.Binary('CSV or Excel File', required=True)
    filename = fields.Char('Filename')
    error_file = fields.Binary('Error File', readonly=True)
    error_filename = fields.Char('Error Filename', readonly=True)

    def action_import(self):
        if not self.data_file:
            raise ValidationError('Please upload a CSV or Excel file.')
        if not self.filename:
            raise ValidationError('Filename is missing.')
        
        ext = os.path.splitext(self.filename)[-1].lower()
        skipped_records = []
        batch_size = 100
        current_batch = []
        created_records = 0
        
        try:
            file_content = base64.b64decode(self.data_file)
            if ext == '.csv':
                csvfile = StringIO(file_content.decode('utf-8'))
                reader = csv.DictReader(csvfile)
                rows = list(reader)
            elif ext in ('.xls', '.xlsx'):
                try:
                    import pandas as pd
                except ImportError:
                    raise ValidationError('pandas library is required for Excel import. Please install it.')
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(file_content)
                    tmp.seek(0)
                    wb = openpyxl.load_workbook(tmp.name)
                    sheet = wb.active
                    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                    header_map = {header: idx for idx, header in enumerate(headers)}
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        zone_name = row[header_map['zone']]
                        ward_name = row[header_map['ward']]
                        colony_name = row[header_map['colony']]
                        
                        # Validate required fields
                        if not all([zone_name, ward_name, colony_name, row[header_map['upicno']]]):
                            _logger.error(f"Skipping row due to missing required fields: {row}")
                            continue
                            
                        zone_id = self.env['ddn.zone'].search([('name', 'ilike', zone_name)], limit=1)
                        ward_id = self.env['ddn.ward'].search([('name', 'ilike', ward_name)], limit=1)
                        colony_id = self.env['ddn.colony'].search([('name', 'ilike', colony_name)], limit=1)
                        
                        if not all([zone_id, ward_id, colony_id]):
                            _logger.error(f"Skipping row due to missing related records: Zone={zone_name}, Ward={ward_name}, Colony={colony_name}")
                            continue
                            
                        vals = {
                            'zone_id': zone_id.id,
                            'ward_id': ward_id.id,
                            'colony_id': colony_id.id,
                            'upic_no': row[header_map['upicno']],
                            'property_no': row[header_map['propperty_id']],
                            'unit_no': row[header_map['unit_no']],
                        }
                        rows.append(vals)
            else:
                raise ValidationError('Unsupported file type. Please upload a .csv, .xls, or .xlsx file.')

            # Process rows and check for duplicates
            total_records = len(rows)
            processed_records = 0
            
            _logger.info(f"Starting import of {total_records} records")
            
            for row in rows:
                try:
                    # Check if property with same UPIC number exists
                    existing_property = self.env['ddn.property.info'].search([
                        ('upic_no', '=', row['upic_no'])
                    ], limit=1)
                    
                    if not existing_property:
                        current_batch.append(row)
                        processed_records += 1
                        
                        # Process batch when it reaches batch_size
                        if len(current_batch) >= batch_size:
                            try:
                                # Create records one by one to handle QR code generation
                                created_records = []
                                for record in current_batch:
                                    try:
                                        new_record = self.env['ddn.property.info'].create(record)
                                        created_records.append(new_record)
                                    except Exception as e:
                                        _logger.error(f"Error creating record: {str(e)}")
                                        continue
                                
                                created_records += len(created_records)
                                _logger.info(f'Batch inserted in DB: {processed_records}/{total_records} records processed. Created: {len(created_records)} records')
                                current_batch = []
                            except Exception as e:
                                _logger.error(f"Error creating batch: {str(e)}")
                                raise
                    else:
                        skipped_records.append(row)
                        processed_records += 1
                        _logger.info(f"Skipped duplicate UPIC: {row['upic_no']}")
                except Exception as e:
                    _logger.error(f"Error processing row: {str(e)}")
                    continue

            # Process remaining records in the last batch
            if current_batch:
                try:
                    # Create records one by one to handle QR code generation
                    created_records = []
                    for record in current_batch:
                        try:
                            new_record = self.env['ddn.property.info'].create(record)
                            created_records.append(new_record)
                        except Exception as e:
                            _logger.error(f"Error creating record: {str(e)}")
                            continue
                    
                    created_records += len(created_records)
                    _logger.info(f'Final batch inserted in DB: {processed_records}/{total_records} records processed. Created: {len(created_records)} records')
                except Exception as e:
                    _logger.error(f"Error creating final batch: {str(e)}")
                    raise

            _logger.info(f"Import completed. Total records processed: {processed_records}, Created: {created_records}, Skipped: {len(skipped_records)}")

            # Generate error file if there are skipped records
            if skipped_records:
                error_wb = openpyxl.Workbook()
                error_sheet = error_wb.active
                
                # Write headers
                headers = ['zone', 'ward', 'colony', 'upicno', 'propperty_id', 'unit_no', 'error_reason']
                for col, header in enumerate(headers, 1):
                    error_sheet.cell(row=1, column=col, value=header)
                
                # Write skipped records
                for row_idx, record in enumerate(skipped_records, 2):
                    error_sheet.cell(row=row_idx, column=1, value=record.get('zone_id'))
                    error_sheet.cell(row=row_idx, column=2, value=record.get('ward_id'))
                    error_sheet.cell(row=row_idx, column=3, value=record.get('colony_id'))
                    error_sheet.cell(row=row_idx, column=4, value=record.get('upic_no'))
                    error_sheet.cell(row=row_idx, column=5, value=record.get('property_no'))
                    error_sheet.cell(row=row_idx, column=6, value=record.get('unit_no'))
                    error_sheet.cell(row=row_idx, column=7, value='Duplicate UPIC number')
                
                # Save error file
                error_file_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                error_wb.save(error_file_path.name)
                
                with open(error_file_path.name, 'rb') as f:
                    error_file_content = f.read()
                
                # Update wizard with error file
                self.write({
                    'error_file': base64.b64encode(error_file_content),
                    'error_filename': 'import_errors.xlsx'
                })
                
                # Clean up temporary file
                os.unlink(error_file_path.name)
                
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'property.import.wizard',
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                }

        except Exception as e:
            _logger.error(f"Import failed: {str(e)}")
            raise ValidationError('Import failed: %s' % str(e))
        
        return {'type': 'ir.actions.act_window_close'}

    def action_test_record(self):
        zone_id = self.env['ddn.zone'].search([('name', '=', 'TEST ZONE')], limit=1)
        ward_id = self.env['ddn.ward'].search([('name', '=', 'TEST WARD')], limit=1)
        colony_id = self.env['ddn.colony'].search([('name', '=', 'TEST COLONY')], limit=1)
        
        if not all([zone_id, ward_id, colony_id]):
            raise ValidationError('TEST records not found in the database.')
        
        self.env['ddn.property.info'].create({
            'zone_id': zone_id.id,
            'ward_id': ward_id.id,
            'colony_id': colony_id.id,
            'upic_no': 'TEST001',
            'property_no': 'TEST001',
            'unit_no': '1'
        })
        
        return {'type': 'ir.actions.act_window_close'} 